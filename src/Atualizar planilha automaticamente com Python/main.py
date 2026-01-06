import openpyxl
import os

dados_xlsx = r"C:/Users/rafae/OneDrive/Área de Trabalho/Atualizar planilha automaticamente com Python/dados.xlsx"

def atualizar_planilha(nome, idade, cargo):
    # Verifica se o arquivo já existe
    if not os.path.exists(dados_xlsx):
        # Criar nova planilha com cabeçalhos
        workbook = openpyxl.Workbook()
        aba = workbook.active
        aba.title = "Dados"
        aba.append(["Nome", "Idade", "Cargo"])  # Cabeçalho
        print("Arquivo não encontrado, criando novo arquivo...")
    else:
        # Carregar planilha existente
        workbook = openpyxl.load_workbook(dados_xlsx)
        aba = workbook.active

    # Verifica duplicidade de dados
    for linha in aba.iter_rows(min_row=2, values_only=True):
        if linha == (nome, idade, cargo):
            print(f"Registro duplicado! Não adicionado: {linha}")
            workbook.close()
            return
    
    # Adicionar nova linha de dados
    aba.append([nome, idade, cargo])
    workbook.save(dados_xlsx)
    print("Planilha atualizada com sucesso!")


funcionarios = [
    ("Rafael", 34, "Aux. Administrativo Industrial"),
    ("Ana", 28, "Analista Financeira"),
    ("Carlos", 41, "Supervisor de Produção"),
    ("Julia", 25, "Assistente de RH"),
    ("Marcos", 38, "Engenheiro de Segurança")
]

for nome, idade, cargo in funcionarios:
    atualizar_planilha(nome, idade, cargo)
